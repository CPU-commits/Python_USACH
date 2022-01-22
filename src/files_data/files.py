import os
import pathlib
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape


from get_into_data import groups
from get_into_data import grades
from get_into_data import forms


wb = Workbook()


def __validate_dir_existence(dirname):
    dir = f'src/{dirname}'
    path = os.path.join(pathlib.Path().resolve(), dir)
    if os.path.exists(path):
        return True
    return False


def __create_dir(dirname):
    dir = f'src/{dirname}'
    path = os.path.join(pathlib.Path().resolve(), dir)
    if os.path.exists(path) is False:
        os.mkdir(path)


def __create_file(dirname, text):
    f = open(dirname, 'w+')
    f.write(text)
    f.close()
    return f


def set_db():
    files = [
        ('forms.json', '[]'),
        ('grades.json', '{}'),
        ('groups.json', '[]'),
        ('answers.json', '{}'),
        ('reminder.json', '[]'),
        ('historical.json', '[]'),
        ('maps.json', '[]'),
    ]
    db_existence = __validate_dir_existence('db')
    if db_existence is False:
        __create_dir('db')
    for file, init_text in files:
        file_existence = __validate_dir_existence(f"db/{file}")
        if file_existence is False:
            __create_file(f"src/db/{file}", init_text)
    __create_dir('output')
    __create_dir('input')


def create_excel_file():
    __create_dir('output')
    __create_dir('input')
    dest_filename = 'src/output/grades.xlsx'
    ws = wb.active
    ws.title = "Grades"
    ws.cell(column=1, row=1, value="Curso (id)")
    ws.cell(column=2, row=1, value="Alumno (id)")
    ws.cell(column=3, row=1, value="Calificación/Nota")
    wb.save(filename=dest_filename)


def __validators(value, col, row, groups_data, course=""):
    if col == 1:
        if value not in groups_data:
            return (True, f"Error en A{row}: Este curso no existe")
    elif col == 2:
        if value not in groups_data[course]:
            return (
                True,
                f"Error en B{row}: Alumno, no existe"
                ", o no es del curso dado"
            )
    return (False, "")


def __validators_group(value, col, row, groups_data, course=""):
    if course != "":
        general_validators, message = __validators(
            value,
            col,
            row,
            groups_data,
            course
        )
    else:
        general_validators, message = __validators(
            value,
            col,
            row,
            groups_data
        )
    if general_validators is False:
        return (False, message)
    if col == 3:
        if grades.validate_grade(str(value)) is False:
            error = f"Error en C{row}: Calificación no tiene formato"
            "'Entero.Decimal'"
            return (
                True,
                error
            )
    return (False, "")


def __format_data(groups_data):
    new_groups = {}
    for groups in groups_data:
        key = groups['id']
        new_groups[key] = list(map(
            lambda student: student['id'],
            groups['students'],
        ))
    return new_groups


def read_excel_file():
    # Validar archivo excel
    excel_file = __validate_dir_existence('input/grades.xlsx')
    if excel_file:
        # Abrir archivo excel
        wb = load_workbook(filename="src/input/grades.xlsx")
        grades_sheet = wb['Grades']
        dimensions = (grades_sheet.calculate_dimension()).split(':')
        if 'C' in dimensions[1]:
            groups_data = groups.get_data()
            format_data = __format_data(groups_data)
            new_grades_data = {}
            large = int(dimensions[1][1])
            general_error = False
            for row in range(2, large+1):
                for col in range(1, 4):
                    value = grades_sheet[
                        f"{get_column_letter(col)}{row}"
                    ].value
                    if col != 2:
                        error, message = __validators_group(
                            value,
                            col,
                            row,
                            format_data
                        )
                    else:
                        error, message = __validators_group(
                            value,
                            col,
                            row,
                            format_data,
                            grades_sheet[f"A{row}"].value
                        )
                    if error:
                        print(message)
                        general_error = True
                    elif col == 3 and general_error is False:
                        group = grades_sheet[f"A{row}"].value
                        student = grades_sheet[f"B{row}"].value
                        if group not in new_grades_data:
                            new_grades_data[group] = {}
                        if student not in new_grades_data[group]:
                            new_grades_data[group][student] = [value]
                        else:
                            new_grades_data[group][student].append(value)
            if general_error is False:
                return {
                    'success': True,
                    'data': new_grades_data
                }
        else:
            print('Deben existir máximo 3 columnas')
    else:
        print('No existe el archivo "grades.xlsx" en input')
    return {
        'success': False
    }


def __validators_form(
    value,
    col,
    row,
    groups_data,
    forms_data,
    sheet,
    course=""
):
    if course != "":
        general_validators, message = __validators(
            value,
            col,
            row,
            groups_data,
            course
        )
    else:
        general_validators, message = __validators(
            value,
            col,
            row,
            groups_data
        )
    if general_validators is False:
        return (False, message)
    if col == 3:
        if int(value) > len(forms_data[sheet]) - 1:
            return (
                False,
                f"Error en C{row}: Esta pregunta no se encuentra"
                "en ninguna de las preguntas almacenadas"
            )
    elif col == 4:
        message = f"Error en D{row}: La respuesta debe ser en escala"
        "de 0 - 10 en enteros"
        if len(value) == 2:
            if value != "10":
                return (False, message)
        elif len(value) > 2:
            return (False, message)
        elif value not in "012345679":
            return (False, message)
    return (True, "")


def __format_data_forms(forms_data):
    new_forms = {}
    for form in forms_data:
        new_forms[form['id']] = form['questions']
    return new_forms


def __get_sheet_key(name, forms_data):
    sheet_key = list(filter(
        lambda x: x['name'] == name,
        forms_data)
    )
    return sheet_key


def read_excel_file_forms():
    excel_file = __validate_dir_existence('input/forms.xlsx')
    general_error = False
    if excel_file:
        wb = load_workbook(filename="src/input/forms.xlsx")
        forms_data = forms.get_data()
        format_data_form = __format_data_forms(
            forms_data
        )
        for sheet_name in wb.sheetnames:
            sheet_key = __get_sheet_key(sheet_name, forms_data)
            if len(sheet_key) == 0:
                print(
                    f"El nombre de formulario: {sheet_name}."
                    "No existe dentro de los formularios"
                )
                general_error = True
                break
            sheet_key = sheet_key[0]['id']
            sheet = wb[sheet_name]
            dimensions = (sheet.calculate_dimension()).split(':')
            large = int(dimensions[1][1])
            if 'D' in dimensions[1] and large > 0:
                groups_data = groups.get_data()
                format_data = __format_data(groups_data)
                new_forms_data = {}
                for row in range(1, large+1):
                    for col in range(1, 5):
                        value = sheet[f"{get_column_letter(col)}{row}"].value
                        if col != 2:
                            error, message = __validators_form(
                                value,
                                col,
                                row,
                                format_data,
                                format_data_form,
                                sheet_key
                            )
                        else:
                            error, message = __validators_form(
                                value,
                                col,
                                row,
                                format_data,
                                format_data_form,
                                sheet_key,
                                sheet[f"A{row}"].value
                            )
                        if error:
                            print(message)
                            general_error = True
                        elif col == 4 and general_error is False:
                            group = sheet[f"A{row}"].value
                            student = sheet[f"B{row}"].value
                            question = sheet[f"C{row}"].value
                            if sheet_key not in new_forms_data:
                                new_forms_data[sheet_key] = {}
                            student_response = {
                                "student": student,
                                "answer": value,
                                "question": question
                            }
                            v = new_forms_data[sheet_key]
                            if group not in v:
                                v[group] = [student_response]
                            else:
                                v[group].append(student_response)
                if general_error is False:
                    return {
                        'success': True,
                        'data': new_forms_data
                    }
            else:
                print('Deben existir máximo 4 columnas, y mínimo una fila')
    else:
        print('No existe el archivo "forms.xlsx" en input')
    return {
        'success': False
    }


def generate_pdf(name_file, title):
    w, h = landscape(letter)
    pdf_file = canvas.Canvas(
        f'src/output/{name_file}.pdf',
        pagesize=landscape(letter)
    )
    title_len = pdf_file.stringWidth(title, 'Times-Roman', 20)
    pdf_file.setFont('Times-Roman', 20)
    pdf_file.drawString((w-title_len)/2, h-50, title)
    return pdf_file, w, h
